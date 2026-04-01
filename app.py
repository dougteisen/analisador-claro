import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ── Google Vision (OCR para PDFs de imagem) ──────────────────────────────────
try:
    from google.cloud import vision
    from google.oauth2 import service_account
    _VISION_DISPONIVEL = True
except ImportError:
    _VISION_DISPONIVEL = False

# ── Google Gemini (Vision + extração inteligente — gratuito) ─────────────────
try:
    import google.generativeai as genai
    _GEMINI_DISPONIVEL = True
except ImportError:
    _GEMINI_DISPONIVEL = False

# ── pdf2image + poppler ───────────────────────────────────────────────────────
try:
    from pdf2image import convert_from_bytes as _pdf2image_convert
    _PDF2IMAGE_DISPONIVEL = True
except ImportError:
    _PDF2IMAGE_DISPONIVEL = False

# FIX 1: st.set_page_config() DEVE ser a primeira chamada Streamlit — sem nada antes
st.set_page_config(
    layout="wide",
    page_title="Target Telecom · Análise de Faturas",
    page_icon="📡"
)

# ===== CSS REDESIGN =====
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

/* ── Base & Background ── */
html, body, [data-testid="stAppViewContainer"] {
    background: #060d1a !important;
}
.main {
    background: #060d1a !important;
}
[data-testid="stAppViewContainer"]::before {
    content: '';
    position: fixed;
    top: -20%;
    left: -10%;
    width: 55%;
    height: 55%;
    background: radial-gradient(ellipse, rgba(16,185,129,0.07) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
}
[data-testid="stAppViewContainer"]::after {
    content: '';
    position: fixed;
    bottom: -10%;
    right: -5%;
    width: 40%;
    height: 40%;
    background: radial-gradient(ellipse, rgba(59,130,246,0.06) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
}
.block-container {
    padding-top: 2rem !important;
    padding-bottom: 3rem !important;
    max-width: 1400px !important;
}

/* ── Tipografia global ── */
*, h1, h2, h3, h4, p, span, div, label {
    font-family: 'DM Sans', sans-serif !important;
    color: #e2e8f0;
}

/* ── Header personalizado ── */
.tt-header {
    display: flex;
    align-items: center;
    gap: 20px;
    padding: 28px 36px;
    background: linear-gradient(135deg, #0d1f2d 0%, #0a1628 100%);
    border: 1px solid rgba(16,185,129,0.18);
    border-radius: 20px;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
}
.tt-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, #10b981, #3b82f6, transparent);
}
.tt-logo-placeholder {
    width: 56px; height: 56px;
    background: linear-gradient(135deg, #10b981, #059669);
    border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    font-size: 26px;
    flex-shrink: 0;
    box-shadow: 0 0 24px rgba(16,185,129,0.35);
}
.tt-brand h1 {
    font-family: 'Syne', sans-serif !important;
    font-size: 1.75rem !important;
    font-weight: 800 !important;
    letter-spacing: -0.02em;
    margin: 0 !important; padding: 0 !important;
    background: linear-gradient(90deg, #ffffff, #a7f3d0);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    line-height: 1.1 !important;
}
.tt-brand p {
    font-size: 0.82rem;
    color: #64748b !important;
    margin: 4px 0 0 0;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    font-weight: 500;
}
.tt-badge {
    margin-left: auto;
    padding: 6px 14px;
    background: rgba(16,185,129,0.1);
    border: 1px solid rgba(16,185,129,0.25);
    border-radius: 20px;
    font-size: 0.72rem;
    color: #10b981 !important;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    font-weight: 600;
}

/* ── Divider ── */
hr {
    border: none !important;
    border-top: 1px solid rgba(255,255,255,0.06) !important;
    margin: 1.5rem 0 !important;
}

/* ── Upload Area ── */
[data-testid="stFileUploader"] {
    background: linear-gradient(135deg, #0d1f2d 0%, #091523 100%) !important;
    border: 1.5px dashed rgba(16,185,129,0.35) !important;
    border-radius: 18px !important;
    padding: 8px 16px !important;
    transition: all 0.3s ease !important;
    position: relative;
}
[data-testid="stFileUploader"]:hover {
    border-color: rgba(16,185,129,0.75) !important;
    background: linear-gradient(135deg, #0f2535 0%, #0c1d2e 100%) !important;
    box-shadow: 0 0 30px rgba(16,185,129,0.08) !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: transparent !important;
    border: none !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] p,
[data-testid="stFileUploaderDropzoneInstructions"] span {
    color: #94a3b8 !important;
    font-size: 0.9rem !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stFileUploaderDropzone"] svg {
    fill: #10b981 !important;
    opacity: 0.8;
}
[data-testid="stFileUploader"] label {
    font-size: 0.95rem !important;
    color: #cbd5e1 !important;
    font-weight: 500 !important;
}

/* ── Métricas ── */
[data-testid="metric-container"] {
    background: linear-gradient(145deg, #0d1f2d, #0a1929) !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 16px !important;
    padding: 20px 24px !important;
    position: relative;
    overflow: hidden;
    transition: transform 0.2s ease, border-color 0.2s ease;
}
[data-testid="metric-container"]:hover {
    transform: translateY(-2px);
    border-color: rgba(16,185,129,0.25) !important;
}
[data-testid="metric-container"]::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #10b981, #3b82f6);
    opacity: 0.6;
}
[data-testid="stMetricLabel"] {
    font-size: 0.72rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    color: #64748b !important;
    font-weight: 600 !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Syne', sans-serif !important;
    font-size: 2rem !important;
    font-weight: 700 !important;
    color: #f1f5f9 !important;
    line-height: 1.2 !important;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 16px !important;
    overflow: hidden !important;
    background: #0a1628 !important;
}
[data-testid="stDataFrame"] table {
    background: transparent !important;
}
[data-testid="stDataFrame"] thead th {
    background: #0d1f2d !important;
    color: #94a3b8 !important;
    font-size: 0.72rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    border-bottom: 1px solid rgba(255,255,255,0.08) !important;
    padding: 12px 16px !important;
}
[data-testid="stDataFrame"] tbody tr:hover td {
    background: rgba(16,185,129,0.05) !important;
}

/* ── Botão de Download ── */
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 12px !important;
    height: 52px !important;
    font-size: 0.9rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
    width: 100% !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 4px 20px rgba(16,185,129,0.25) !important;
}
[data-testid="stDownloadButton"] > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(16,185,129,0.4) !important;
    background: linear-gradient(135deg, #047857 0%, #059669 100%) !important;
}

/* ── Spinner e Progress ── */
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #10b981, #3b82f6) !important;
    border-radius: 4px !important;
}
[data-testid="stProgress"] {
    background: rgba(255,255,255,0.06) !important;
    border-radius: 4px !important;
}

/* ── Alerts / Errors ── */
[data-testid="stAlert"] {
    border-radius: 12px !important;
    border-left: 3px solid #ef4444 !important;
    background: rgba(239,68,68,0.08) !important;
}

/* ── Sidebar (se abrir) ── */
[data-testid="stSidebar"] {
    background: #060d1a !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
}

/* ── Seção de resultados ── */
.tt-section-title {
    font-family: 'Syne', sans-serif !important;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #475569 !important;
    margin: 8px 0 16px 2px;
}
.tt-divider {
    height: 1px;
    background: linear-gradient(90deg, rgba(16,185,129,0.3), rgba(59,130,246,0.15), transparent);
    border: none;
    margin: 24px 0;
}
</style>
""", unsafe_allow_html=True)

# ===== HEADER =====
if os.path.exists("logo.png"):
    col1, col2 = st.columns([1, 5])
    with col1:
        st.image("logo.png", width=180)
    with col2:
        st.markdown("""
        <div style="padding: 12px 0;">
            <div style="font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;
                        background:linear-gradient(90deg,#fff,#a7f3d0);
                        -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                        background-clip:text;letter-spacing:-0.02em;">TARGET TELECOM</div>
            <div style="font-size:0.78rem;color:#64748b;text-transform:uppercase;
                        letter-spacing:0.1em;font-weight:500;margin-top:4px;">
                Inteligência em Faturas Corporativas
            </div>
        </div>
        """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div class="tt-header">
        <div class="tt-logo-placeholder">📡</div>
        <div class="tt-brand">
            <h1>TARGET TELECOM</h1>
            <p>Inteligência em Faturas Corporativas</p>
        </div>
        <div class="tt-badge">✦ Análise Automática</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div class="tt-divider"></div>', unsafe_allow_html=True)

# ===== UPLOAD =====
st.markdown('<p class="tt-section-title">📂 Carregar Fatura</p>', unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Arraste o PDF da fatura ou clique para selecionar — múltiplos arquivos suportados",
    type="pdf",
    accept_multiple_files=True,
    label_visibility="visible"
)

# ===== GOOGLE VISION — cliente único (singleton) =====
# FIX 2: instanciado UMA vez aqui, não repetido dentro de cada função
_vision_client = None

def _get_vision_client():
    """Retorna o cliente Vision, criando-o na primeira chamada."""
    global _vision_client
    if _vision_client is not None:
        return _vision_client
    if not _VISION_DISPONIVEL:
        return None
    try:
        creds_dict = st.secrets["GOOGLE_CREDENTIALS"]
        credentials = service_account.Credentials.from_service_account_info(creds_dict)
        _vision_client = vision.ImageAnnotatorClient(credentials=credentials)
        return _vision_client
    except Exception as e:
        st.warning(f"⚠️ Google Vision não configurado: {e}. PDFs de imagem não serão suportados.")
        return None

def extrair_texto_com_ocr(img_bytes: bytes) -> str:
    """
    Envia imagem PNG (bytes) para Google Vision e retorna o texto extraído.
    FIX 6: recebe bytes puros (não BytesIO) — chamador deve usar .getvalue()
    FIX 2: usa cliente singleton, não recria a cada chamada
    """
    client = _get_vision_client()
    if client is None:
        return ""
    try:
        image = vision.Image(content=img_bytes)
        response = client.document_text_detection(image=image)
        if response.error.message:
            return ""
        return response.full_text_annotation.text or ""
    except Exception:
        return ""

# ===== UTILITÁRIOS =====

def normalizar_numero(num_str: str) -> str:
    """Remove parênteses e espaços do número de telefone."""
    return num_str.replace("(", "").replace(")", "").replace(" ", "")

def extrair_cliente(texto: str) -> str:
    linhas = texto.split("\n")
    for i, linha in enumerate(linhas):
        if "nº do cliente" in linha.lower():
            if i >= 1:
                nome = linhas[i - 1].strip().upper()
                nome = re.sub(r'[\\/:*?"<>|]', "", nome)
                nome = re.sub(r"\s+", " ", nome)
                return nome
    return "CLIENTE"

def extrair_vencimento(texto: str) -> str:
    # FIX #5: usa re.DOTALL para capturar mesmo com quebra de linha
    match = re.search(r"Nº da conta:.*?(\d{2}/\d{2}/\d{4})", texto, re.DOTALL)
    if match:
        return match.group(1)
    # fallback: busca a data de vencimento diretamente
    match = re.search(r"Vencimento\s*\n\s*(\d{2}/\d{2}/\d{4})", texto)
    if match:
        return match.group(1)
    return ""

def normalizar_para_comparacao(texto: str) -> str:
    """
    Normaliza texto removendo acentos para comparações robustas.
    Útil para lidar com variações de OCR em texto português.
    """
    import unicodedata
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII').upper()

def extrair_linhas(texto: str) -> list:
    # Padrão principal: (11) 98936 0484
    linhas = re.findall(r"\(\d{2}\)\s\d{5}\s\d{4}", texto)
    # Fallback OCR: número pode vir sem parênteses ou com espaçamento diferente
    if not linhas:
        linhas = re.findall(r"\d{2}\s\d{5}\s\d{4}", texto)
    lista = []
    for l in linhas:
        num = normalizar_numero(l)
        if num not in lista:
            lista.append(num)
    return lista

# Padrão de divisão de blocos — tolerante a variações de acentuação do OCR
# Cobre: LIGAÇÕES/LIGACOES, SERVIÇOS/SERVICOS, com ou sem acento
_PADRAO_DIVISOR_BLOCO = re.compile(
    r"DETALHAMENTO\s+DE\s+LIGA[CÇ][OÕ0]ES\s+E\s+SERVI[CÇ]OS\s+DO\s+CELULAR",
    re.IGNORECASE
)

def extrair_blocos_por_linha(texto: str) -> dict:
    """
    Divide o texto em blocos por linha telefônica.
    Tolerante a variações de acentuação do OCR (LIGACOES vs LIGAÇÕES, etc).
    """
    # Estratégia 1: split tolerante a acentos (cobre OCR e PDF digital)
    blocos = _PADRAO_DIVISOR_BLOCO.split(texto)

    # Fallback: se o split não funcionou, tenta com texto normalizado (sem acentos)
    if len(blocos) <= 1:
        texto_norm = normalizar_para_comparacao(texto)
        posicoes = [m.start() for m in re.finditer(
            r"DETALHAMENTO\s+DE\s+LIGACOES\s+E\s+SERVICOS\s+DO\s+CELULAR", texto_norm
        )]
        if posicoes:
            blocos = [texto[s:e] for s, e in zip(
                [0] + posicoes,
                posicoes + [len(texto)]
            )]

    resultado = {}
    for bloco in blocos:
        # Padrão principal com parênteses
        num = re.search(r"\(\d{2}\)\s*\d{5}\s*\d{4}", bloco)
        if num:
            chave = normalizar_numero(num.group(0))
            resultado[chave] = bloco
        else:
            # Fallback: número sem parênteses (variação OCR)
            num = re.search(r"^\s*\d{2}\s+\d{5}\s+\d{4}", bloco, re.MULTILINE)
            if num:
                chave = normalizar_numero(num.group(0))
                resultado[chave] = bloco
    return resultado

def extrair_mensalidades(blocos: dict) -> dict:
    """
    Captura o TOTAL de cada linha com múltiplos fallbacks para suportar
    variações de formatação em PDFs digitais e PDFs de imagem (OCR).
    """
    mapa = {}
    for linha, bloco in blocos.items():
        total_pdf = 0.0
        total_str = None

        # Estratégia 1: "TOTAL R$59,15" ou "TOTAL R$ 59,15" (com ou sem espaço)
        m = re.search(r"TOTAL\s*R\$\s*([\d\.,]+)", bloco)
        if m:
            total_str = m.group(1)
            try:
                total_pdf = float(total_str.replace(".", "").replace(",", "."))
            except (ValueError, TypeError):
                pass

        # Estratégia 2: "TOTAL RS59,15" — OCR confunde $ com S
        if total_pdf == 0.0:
            m = re.search(r"TOTAL\s+R[S$]\s*([\d\.,]+)", bloco, re.IGNORECASE)
            if m:
                total_str = m.group(1)
                try:
                    total_pdf = float(total_str.replace(".", "").replace(",", "."))
                except (ValueError, TypeError):
                    pass

        # Estratégia 3: somar valores positivos da seção de mensalidades
        # — fallback robusto quando TOTAL não é detectado e também
        #   serve para ignorar descontos negativos (ex: Desconto Dados GPRS)
        secao = re.search(r"Mensalidades e Pacotes Promocionais(.*?)TOTAL", bloco, re.DOTALL)
        soma_positivos = 0.0
        if secao:
            for lb in secao.group(1).split("\n"):
                mv = re.search(r"(-?[\d]+,\d{2})$", lb.strip())
                if mv:
                    try:
                        val = float(mv.group(1).replace(".", "").replace(",", "."))
                        if val > 0:
                            soma_positivos += val
                    except (ValueError, TypeError):
                        pass

        # Se descontos reduziram o total, ou se total não foi detectado, usa soma
        if soma_positivos > total_pdf + 0.01:
            mapa[linha] = f"{soma_positivos:.2f}".replace(".", ",")
        elif total_str:
            mapa[linha] = total_str

    return mapa

def extrair_pacote_e_passaporte(blocos: dict) -> dict:
    resultado = {}
    for linha, bloco in blocos.items():
        pacote = "-"
        passaporte = "-"
        valor_passaporte = "0"

        # Captura todos os planos possíveis: Claro Pós/Life Ilimitado/Controle e Plano Wi-Fi
        m = re.search(
            r"((?:Claro\s+(?:Pós|Life Ilimitado|Controle)|Plano de Internet Wi-Fi)\s+\d+\s*GB)",
            bloco
        )
        if m:
            pacote = m.group(1).strip()

        # Busca seção de mensalidades e pacotes
        secao = re.search(
            r"Mensalidades e Pacotes Promocionais(.*?)TOTAL\s+R\$",
            bloco,
            re.DOTALL
        )
        if secao:
            trecho = secao.group(1)
            for linha_bloco in trecho.split("\n"):
                linha_limpa = linha_bloco.strip()
                if "Claro Passaporte" not in linha_limpa:
                    continue
                # Captura nome do passaporte e valor no final da linha
                m = re.search(r"(Claro Passaporte.*?GB).*?([\d]+,\d{2})$", linha_limpa)
                if m:
                    passaporte = m.group(1).strip()
                    valor_passaporte = m.group(2).strip()
                    break

        resultado[linha] = {
            "Pacote": pacote,
            "Passaporte": passaporte,
            "Valor Passaporte": valor_passaporte
        }
    return resultado

def extrair_detalhamento(blocos: dict) -> dict:
    mapa = {}
    for linha, bloco in blocos.items():
        internet = "0"

        # Estratégia 1: âncora em "Serviços (Torpedos" + linha "Internet X"
        # Funciona em PDFs digitais onde a seção está bem estruturada
        m = re.search(
            r"Servi[çc]os\s*\(Torpedos.*?^Internet\s+([\d\.]+[,\.][\d]+)\s+0[,\.]00",
            bloco, re.DOTALL | re.MULTILINE | re.IGNORECASE
        )
        if m:
            internet = m.group(1)
        else:
            # Estratégia 2: linha "Internet X,XXX 0,00" sem exigir âncora
            # Mais tolerante a variações de OCR
            m = re.search(
                r"^Internet\s+([\d\.]+[,\.][\d]+)\s+0[,\.]00",
                bloco, re.MULTILINE | re.IGNORECASE
            )
            if m:
                internet = m.group(1)
            else:
                # Estratégia 3: Subtotal após seção Internet (ignora Subtotal 0,00 do roaming)
                m = re.search(
                    r"Internet\s+[\d\.,]+.*?Subtotal\s+([\d\.]+[,\.][\d]+)\s+0[,\.]00",
                    bloco, re.DOTALL | re.IGNORECASE
                )
                if m:
                    try:
                        val = float(m.group(1).replace(".", "").replace(",", "."))
                        if val > 0:
                            internet = m.group(1)
                    except (ValueError, TypeError):
                        pass

        # Minutos: "Xmin Ys", "Xmin", "Xs" — tolerante a variações OCR
        minutos = "0"
        m = re.search(r"TOTAL\s+([\d]+min[\d]*s?|[\d]+s)", bloco)
        if m:
            minutos = m.group(1)

        mapa[linha] = {
            "Internet (MB)": internet,
            "Minutos": minutos
        }
    return mapa

def to_float(valor) -> float:
    try:
        return float(str(valor).replace(".", "").replace(",", "."))
    except (ValueError, TypeError):
        return 0.0

def extrair_gb_pacote(pacote: str) -> int:
    m = re.search(r"(\d+)\s*GB", str(pacote))
    return int(m.group(1)) if m else 0

# ===== PLANO COMPARTILHADO =================================================
# Funções 100% separadas do fluxo individual.
# O código existente para planos individuais NÃO é tocado.

def _detectar_plano_compartilhado(texto: str) -> bool:
    """Retorna True se a fatura for de plano compartilhado."""
    return bool(re.search(r"Claro Total Compartilhado|Total Compartilhado", texto, re.IGNORECASE))

def _extrair_plano_compartilhado_info(texto: str) -> dict:
    """
    Extrai nome e valor do plano compartilhado da capa.
    Ex: "Claro Total Compartilhado 500GB" → {"nome": "Claro Total Compartilhado 500GB", "valor": 695.48}
    """
    # Busca padrão: "Oferta Conjunta Claro MIX  695,48" seguido de "Claro Total Compartilhado XGB"
    m_valor = re.search(
        r"Oferta Conjunta Claro MIX\s+([\d\.]+,\d{2})\s*\n\s*(Claro Total Compartilhado[^\n\[]+)",
        texto
    )
    if m_valor:
        valor_str = m_valor.group(1)
        nome = m_valor.group(2).strip()
        return {"nome": nome, "valor": to_float(valor_str)}

    # Fallback: busca direta pelo nome e valor na seção plano contratado
    m_nome = re.search(r"(Claro Total Compartilhado[^\n\[]+)", texto)
    m_val2 = re.search(r"Oferta Conjunta Claro MIX\s+([\d\.]+,\d{2})", texto)
    if m_nome and m_val2:
        return {"nome": m_nome.group(1).strip(), "valor": to_float(m_val2.group(1))}

    return {"nome": "Claro Total Compartilhado", "valor": 0.0}

def _extrair_blocos_compartilhado(texto: str) -> dict:
    """
    Extrai blocos de detalhamento por linha para plano compartilhado.
    Retorna dict: {numero: texto_do_bloco}
    """
    padrao = r"DETALHAMENTO DE LIGAÇÕES E SERVIÇOS DO CELULAR \((\d{2})\)\s*(\d{4,5}\s*\d{4})"
    posicoes = [(m.start(), normalizar_numero(m.group(1) + m.group(2).replace(" ", "")))
                for m in re.finditer(padrao, texto)]

    blocos = {}
    for idx, (pos, num) in enumerate(posicoes):
        fim = posicoes[idx + 1][0] if idx + 1 < len(posicoes) else len(texto)
        blocos[num] = texto[pos:fim]
    return blocos

def _extrair_internet_compartilhado(bloco: str) -> float:
    """Extrai MB do Subtotal de internet do bloco de uma linha."""
    # Subtotal
    m = re.search(r"Subtotal\s+([\d\.]+,\d{3})", bloco)
    if m:
        return to_float(m.group(1))
    # Só Internet + meses anteriores
    vals = re.findall(r"Internet(?:\s*[-–]\s*meses\s+anteriores)?\s+([\d\.]+,\d{3})", bloco)
    return sum(to_float(v) for v in vals)

def _extrair_minutos_compartilhado(bloco: str) -> str:
    """Extrai duração total de ligações do bloco."""
    m = re.search(r"TOTAL\s+(\d+min(?:\d+s)?|\d+s)", bloco)
    if m:
        return m.group(1)
    return "0"

def _extrair_passaporte_compartilhado(bloco: str) -> tuple:
    """
    Extrai passaporte e valor do bloco de uma linha.
    Retorna (nome_passaporte, valor_float)
    """
    # Busca dentro da seção de Mensalidades apenas (antes de "Ligações" ou "Serviços")
    secao_mens = re.split(r"Ligações|Serviços \(Torpedos|TOTAL\s+R\$", bloco)[0]
    m = re.search(
        r"(Claro Passaporte[^\n]{3,50}?)\s+([\d]{1,3},\d{2})\s*$",
        secao_mens, re.MULTILINE
    )
    if m:
        nome = m.group(1).strip()
        # Garante que não capturou lixo da NF (nome deve ter menos de 50 chars e sem dígitos demais)
        if len(nome) < 60 and not re.search(r"\d{4}", nome):
            return nome, to_float(m.group(2))
    return "-", 0.0

def _extrair_linhas_compartilhado_capa(texto: str) -> list:
    """
    Extrai todos os números de telefone das tabelas de cobranças por celular na capa.
    Ex: "(11) 94463 9555  (11) 96326 8240 ..."
    """
    numeros = set()
    for m in re.finditer(r"\((\d{2})\)\s*(\d{4,5}\s*\d{4})", texto):
        num = normalizar_numero(m.group(1) + m.group(2).replace(" ", ""))
        numeros.add(num)
    return sorted(numeros)

def processar_pdf_compartilhado(texto: str, cliente: str, vencimento: str) -> tuple:
    """
    Processa fatura de plano compartilhado a partir do texto extraído.
    Retorna (DataFrame, cliente, vencimento) no mesmo formato que processar_pdf.
    """
    plano = _extrair_plano_compartilhado_info(texto)
    nome_plano = plano["nome"]
    valor_plano = plano["valor"]

    blocos = _extrair_blocos_compartilhado(texto)

    # Todos os números da fatura (via detalhamento)
    todos_numeros = list(blocos.keys())
    if not todos_numeros:
        # Fallback: extrai da capa
        todos_numeros = _extrair_linhas_compartilhado_capa(texto)

    dados = []
    for num in todos_numeros:
        bloco = blocos.get(num, "")
        internet_mb  = _extrair_internet_compartilhado(bloco)
        minutos      = _extrair_minutos_compartilhado(bloco)
        pass_nome, pass_val = _extrair_passaporte_compartilhado(bloco)

        # Mensalidade individual (assinatura)
        mensalidade_ind = 0.0
        m_mens = re.search(r"Oferta Claro Total Mix Plugin Smartphone\s+([\d]+,\d{2})", bloco)
        if m_mens:
            mensalidade_ind = to_float(m_mens.group(1))

        total_linha = mensalidade_ind + pass_val

        dados.append({
            "Linha":                  num,
            "Internet (MB)":          internet_mb,
            "Internet (MB) fmt":      _fmt_mb_display(internet_mb),
            "Pacote de dados":        nome_plano,
            "Mensalidade":            f"R$ {mensalidade_ind:.2f}".replace(".", ","),
            "Passaporte":             pass_nome,
            "Mensalidade Passaporte": f"R$ {pass_val:.2f}".replace(".", ",") if pass_val else "-",
            "Total por linha":        f"R$ {total_linha:.2f}".replace(".", ","),
            "Minutos":                minutos,
        })

    df = pd.DataFrame(dados)

    # Classificação de perfil (baseada no consumo individual vs franquia total)
    franquia_total_mb = extrair_gb_pacote(nome_plano) * 1024
    total_consumido   = df["Internet (MB)"].sum() if not df.empty else 0

    def classificar_compartilhado(mb):
        if mb == 0:
            return "⚪ Baixo"
        # Percentual do consumo individual em relação à média esperada
        n = max(len(dados), 1)
        media_esperada = franquia_total_mb / n if franquia_total_mb > 0 else 0
        if media_esperada == 0:
            # Fallback: classifica por GB absoluto
            gb = mb / 1024
            if gb < 1:   return "⚪ Baixo"
            if gb < 5:   return "🟡 Médio"
            return "🔴 Alto"
        ratio = mb / media_esperada
        if ratio < 0.3:  return "⚪ Baixo"
        if ratio < 0.8:  return "🟡 Médio"
        return "🔴 Alto"

    df["Perfil"] = df["Internet (MB)"].apply(classificar_compartilhado)

    def em_uso_compartilhado(row):
        minutos_str = str(row["Minutos"]).strip().lower()
        sem_minutos = re.fullmatch(r"0[^\d]*|", minutos_str) is not None
        if row["Internet (MB)"] == 0 and sem_minutos:
            return "Não"
        return "Sim"

    df["Em Uso"] = df.apply(em_uso_compartilhado, axis=1)

    def estrategia_compartilhado(row):
        if row["Em Uso"] == "Não":
            return "⚪ Manter"
        if "Baixo" in row["Perfil"]:
            return "🟡 Sustentar plano"
        if "Médio" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        if "Alto" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        return ""

    df["Estratégia Comercial"] = df.apply(estrategia_compartilhado, axis=1)

    # Linha extra do plano compartilhado (somada no Total por linha)
    linha_plano = {
        "Linha":                  "PLANO COMPARTILHADO",
        "Internet (MB)":          0.0,
        "Internet (MB) fmt":      "-",
        "Pacote de dados":        nome_plano,
        "Mensalidade":            f"R$ {valor_plano:.2f}".replace(".", ","),
        "Passaporte":             "-",
        "Mensalidade Passaporte": "-",
        "Total por linha":        f"R$ {valor_plano:.2f}".replace(".", ","),
        "Minutos":                "-",
        "Perfil":                 "-",
        "Em Uso":                 "-",
        "Estratégia Comercial":   "-",
    }
    df = pd.concat([df, pd.DataFrame([linha_plano])], ignore_index=True)

    return df, cliente, vencimento



_PROMPT_FATURA = """Você é um extrator especializado em faturas da Claro Empresas (Brasil).

PARTE 1 — METADADOS DA CAPA
Da primeira página extraia:
- "cliente": razão social (ex: "PROXY PRODUTOS ORTOPEDICOS LTDA"). NÃO use número de conta.
- "vencimento": data de vencimento DD/MM/AAAA (ex: "17/04/2026")

PARTE 2 — UMA ENTRADA POR SEÇÃO DE DETALHAMENTO
Cada seção começa com cabeçalho vermelho/colorido:
  "DETALHAMENTO DE LIGAÇÕES E SERVIÇOS DO CELULAR (XX) XXXXX XXXX"

⚠️ REGRA ABSOLUTA: retorne UMA entrada para CADA seção encontrada, SEM EXCEÇÃO.
NUNCA pule uma seção, mesmo que ela não tenha internet, ligações, ou pareça vazia.
Seções de linhas sem uso devem aparecer com internet_mb "0" e minutos "0".
Conte as seções antes de montar o JSON e confirme que o número de entradas bate.

Processe cada seção ISOLADAMENTE. Os dados pertencem EXCLUSIVAMENTE ao número
do cabeçalho daquela seção.

━━━ "linha" ━━━
Número do cabeçalho. 11 dígitos sem espaços/parênteses.
"(11) 93235 6185" → "11932356185"

━━━ "pacote" ━━━
DENTRO desta seção, em "Mensalidades e Pacotes Promocionais":

A estrutura pode variar. Dois exemplos reais desta fatura:

  Exemplo A (Claro Pós 40GB):
    Oferta Conjunta Claro MIX       52,48
      App incluso na oferta – ...     –
      Claro Pós 40GB                  –   ← pacote correto
      Aplicativos Digitais            –

  Exemplo B (Claro Pós 10GB — com Bônus):
    Oferta Conjunta Claro MIX       48,49
      Bônus de Internet Turbo – 4GB  0,00
      Claro Pós 10GB                  –   ← pacote correto
      Pacote Mobilidade              0,00

O pacote é SEMPRE a linha com "GB" no nome após "Claro Pós", "Claro Controle", etc.
Ignore: "Oferta Conjunta", "App incluso", "Bônus de Internet Turbo", "Aplicativos Digitais", "Pacote Mobilidade", "Pacote Redes Sociais".

⚠️ CADA SEÇÃO TEM SEU PRÓPRIO PACOTE — leia dentro desta seção específica.
NÃO copie de outra seção. Se não encontrar linha com GB, use "-".

━━━ "mensalidade_total" ━━━
Linha "TOTAL" da subseção "Mensalidades e Pacotes Promocionais" desta seção.
Formato: "52,48" (vírgula decimal, sem R$).

━━━ "internet_mb" ━━━
DENTRO desta seção, em "Serviços (Torpedos, Hits, Jogos, etc.) → Internet (MB)":

  Serviço                  Mbytes Utilizados   Tarifa   Valor
  Internet                     1.455.161        0,00    0,00
  Internet – meses ant.           60.552        0,00    0,00
  Subtotal                     1.515.713                0,00  ← USE ESTE

REGRAS em ordem de prioridade:
1. Use "Subtotal" se existir
2. Some "Internet" + "Internet – meses anteriores" se não houver Subtotal
3. Use só "Internet" se não houver meses anteriores
4. Retorne "0" se não houver seção Internet nesta seção

⚠️ O Subtotal é SEMPRE maior ou igual à linha "Internet".
⚠️ Nunca use valores da tabela "Detalhes da Internet móvel" (datas diárias).
Retorne como STRING com pontos: "1.515.713", "12.722", "387.484", "0"

━━━ "minutos" ━━━
Linha "TOTAL" do rodapé desta seção, coluna "Duração".
Exemplos: "26min12s", "42s", "1min30s", "0".
⚠️ Cada seção tem seu TOTAL próprio. NÃO copie de outra seção.
Se não houver ligações, retorne "0".

━━━ "passaporte" / "valor_passaporte" ━━━
Se houver "Claro Passaporte" em Mensalidades → nome e valor. Senão "-" e "0".

━━━ SAÍDA ━━━
SOMENTE JSON válido, sem markdown.

{
  "cliente": "PROXY PRODUTOS ORTOPEDICOS LTDA",
  "vencimento": "17/04/2026",
  "linhas": [
    {
      "linha": "11932356185",
      "pacote": "Claro Pós 40GB",
      "mensalidade_total": "52,48",
      "internet_mb": "2.037.312",
      "minutos": "30s",
      "passaporte": "-",
      "valor_passaporte": "0"
    },
    {
      "linha": "11978388723",
      "pacote": "Claro Pós 10GB",
      "mensalidade_total": "48,49",
      "internet_mb": "7.308.804",
      "minutos": "42s",
      "passaporte": "-",
      "valor_passaporte": "0"
    }
  ]
}
"""

_PROMPT_VERIFICAR_INTERNET = """Analise esta fatura Claro Empresas.
Para CADA seção "DETALHAMENTO DE LIGAÇÕES E SERVIÇOS DO CELULAR (XX) XXXXX XXXX":

Encontre DENTRO DESTA SEÇÃO a subseção "Internet (MB)" com esta estrutura:
  Serviço              Mbytes Utilizados
  Internet                 1.455.161
  Internet – meses ant.       60.552
  Subtotal                 1.515.713   ← retorne este valor

Regras:
1. Use "Subtotal" quando existir
2. Se não houver Subtotal: some "Internet" + "Internet – meses anteriores"
3. Se só houver "Internet": use esse valor
4. Se não houver seção Internet: retorne "0"

⚠️ NUNCA use valores da tabela de datas diárias ("Detalhes da Internet móvel").
⚠️ Cada número de linha tem seus próprios valores — não misture.

Retorne SOMENTE JSON:
[
  {"linha": "11932356185", "internet_mb": "2.037.312"},
  {"linha": "11978388723", "internet_mb": "7.308.804"},
  {"linha": "11932356313", "internet_mb": "1.515.713"},
  {"linha": "11945701012", "internet_mb": "12.722"},
  {"linha": "11947961230", "internet_mb": "387.484"},
  {"linha": "11945704141", "internet_mb": "0"},
  {"linha": "11978110855", "internet_mb": "0"}
]

Retorne o Subtotal exatamente como aparece na fatura. Se zero/ausente, retorne "0".
"""


# ── Prompt exclusivo para plano compartilhado em PDF de imagem ───────────────
_PROMPT_FATURA_COMPARTILHADA = """Você é um extrator especializado em faturas da Claro Empresas (Brasil).
Esta fatura é de PLANO COMPARTILHADO. Siga as instruções com máxima atenção.

━━━ PARTE 1 — METADADOS DA CAPA (primeira página) ━━━
- "cliente": razão social da empresa. NÃO use número de conta.
- "vencimento": data de vencimento DD/MM/AAAA.

━━━ PARTE 2 — PLANO COMPARTILHADO (primeira página, seção "1. PLANO CONTRATADO") ━━━

A estrutura REAL desta fatura na capa é:

  1. PLANO CONTRATADO                                          VALOR R$
  Compartilhado
  Oferta Conjunta Claro MIX                                    310,28
    Claro Total Compartilhado 150GB [192]
    Aplicativos Digitais
  Individual
  Oferta Claro Total Mix Plugin Smartphone                     297,80
    Assinatura Smartphone [192]
    Aplicativos Digitais
  SUBTOTAL – PLANO CONTRATADO                          R$      608,08

Extraia:
- "nome_plano_compartilhado": linha que contém "Claro Total Compartilhado" seguida de XGB.
  Ignore o código entre colchetes como [192]. Exemplo: "Claro Total Compartilhado 150GB"
- "valor_plano_compartilhado": valor numérico NA MESMA LINHA de "Oferta Conjunta Claro MIX"
  (coluna direita). Exemplo: "310,28". Formato: vírgula decimal, sem R$.

⚠️ NÃO confunda com o valor "Individual" (Oferta Claro Total Mix Plugin Smartphone).
⚠️ NÃO use o SUBTOTAL PLANO CONTRATADO — esse é o total de tudo.
⚠️ Se o valor aparecer como "310.28" (ponto decimal), converta para "310,28".

━━━ PARTE 3 — UMA ENTRADA POR SEÇÃO DE DETALHAMENTO ━━━
Cada seção começa com cabeçalho colorido/vermelho:
  "DETALHAMENTO DE LIGAÇÕES E SERVIÇOS DO CELULAR (XX) XXXXX XXXX"

⚠️ REGRA ABSOLUTA: retorne UMA entrada para CADA seção, SEM EXCEÇÃO.
NUNCA pule uma seção. Seções sem uso → internet_mb "0" e minutos "0".
Conte as seções antes de montar o JSON e confirme que o número de entradas bate.
Processe cada seção ISOLADAMENTE — dados de uma seção nunca vazam para outra.

━━━ "linha" ━━━
Número do cabeçalho. 11 dígitos sem espaços/parênteses.
"(11) 99946 5790" → "11999465790"

━━━ "mensalidade_individual" ━━━
DENTRO desta seção, em "Mensalidades e Pacotes Promocionais":

Estrutura real:
  Oferta Claro Total Mix Plugin Smartphone    21,37
    Assinatura Smartphone [192]                  –
    Aplicativos Digitais                         –
  TOTAL                                       R$ 21,37

Use o valor do TOTAL desta subseção. Formato: vírgula decimal, sem R$.
⚠️ O valor PODE variar por linha (ex: "19,99" em uma linha, "21,37" nas demais).
⚠️ NUNCA copie de outra seção.

━━━ "internet_mb" ━━━
DENTRO desta seção, em "Serviços (Torpedos, Hits, Jogos, etc.) → Internet (MB)":

Estrutura real:
  Serviço                  Mbytes Utilizados   Tarifa (R$)   Valor Cobrado (R$)
  Internet                    12.916,518           0,00            0,00
  Internet – meses ant.          596,958           0,00            0,00
  Subtotal                    13.513,476                           0,00  ← USE ESTE

Regras em ordem de prioridade:
1. Use o valor da linha "Subtotal" se existir — é sempre a soma correta
2. Se não houver Subtotal: some "Internet" + "Internet – meses anteriores"
3. Se só houver "Internet": use esse valor
4. Retorne "0" se não houver seção Internet

⚠️ ATENÇÃO: os valores de Mbytes têm VÍRGULA como separador decimal e PONTO como milhar.
   Exemplo: "13.513,476" significa treze mil quinhentos e treze vírgula quatrocentos e setenta e seis MB.
   Retorne EXATAMENTE como aparece na fatura: "13.513,476" — NÃO transforme em "13513476".
⚠️ NUNCA use valores da tabela "Detalhes da Internet móvel" (datas diárias com valores pequenos).
⚠️ Nunca some valores de seções diferentes.

━━━ "minutos" ━━━
Linha "TOTAL" do RODAPÉ desta seção (última linha da seção), coluna "Duração".
Exemplos: "15min6s", "393min42s", "0".
⚠️ Cada seção tem seu TOTAL próprio. NÃO copie de outra seção.
Se não houver ligações, retorne "0".

━━━ "passaporte" / "valor_passaporte" ━━━
Se houver "Claro Passaporte" em Mensalidades → nome e valor. Senão: "-" e "0".

━━━ SAÍDA ━━━
SOMENTE JSON válido, sem markdown, sem texto extra.

{
  "tipo": "compartilhado",
  "cliente": "CONFRUTY ALIMENTOS EIRELI",
  "vencimento": "24/03/2026",
  "nome_plano_compartilhado": "Claro Total Compartilhado 150GB",
  "valor_plano_compartilhado": "310,28",
  "linhas": [
    {
      "linha": "11919030001",
      "mensalidade_individual": "19,99",
      "internet_mb": "33.511",
      "minutos": "0",
      "passaporte": "-",
      "valor_passaporte": "0"
    },
    {
      "linha": "11989669622",
      "mensalidade_individual": "21,37",
      "internet_mb": "13.513,476",
      "minutos": "11min6s",
      "passaporte": "-",
      "valor_passaporte": "0"
    }
  ]
}
"""

# ── Prompt de verificação de internet para plano compartilhado ───────────────
_PROMPT_VERIFICAR_INTERNET_COMPARTILHADO = """Analise esta fatura Claro Empresas (plano compartilhado).
Para CADA seção "DETALHAMENTO DE LIGAÇÕES E SERVIÇOS DO CELULAR (XX) XXXXX XXXX":

Encontre DENTRO DESTA SEÇÃO EXCLUSIVAMENTE a subseção "Internet (MB)":

  Serviço                  Mbytes Utilizados   Tarifa (R$)   Valor Cobrado (R$)
  Internet                    12.916,518           0,00            0,00
  Internet – meses ant.          596,958           0,00            0,00
  Subtotal                    13.513,476                           0,00  ← USE ESTE

Regras:
1. Use o valor da linha "Subtotal" quando existir
2. Se não houver Subtotal: some "Internet" + "Internet – meses anteriores"
3. Se só houver "Internet": use esse valor
4. Se não houver seção Internet: retorne "0"

⚠️ CRÍTICO: os Mbytes usam PONTO como separador de milhar e VÍRGULA como decimal.
   "13.513,476" = treze mil quinhentos e treze MB. Retorne EXATAMENTE como está na fatura.
⚠️ Cada seção tem seu próprio Subtotal — NUNCA misture valores entre seções diferentes.
⚠️ NUNCA use a tabela "Detalhes da Internet móvel" (contém valores diários pequenos).

Retorne SOMENTE JSON:
[
  {"linha": "11919030001", "internet_mb": "33.511"},
  {"linha": "11989669622", "internet_mb": "13.513,476"}
]
"""


_PROMPT_CAPA_COMPARTILHADA = """Analise SOMENTE a primeira página desta fatura Claro Empresas.

Na seção "1. PLANO CONTRATADO", subseção "Compartilhado", extraia:

1. O nome do plano: linha que contém "Claro Total Compartilhado" seguida de XGB (ex: "Claro Total Compartilhado 150GB"). Ignore códigos como [192].
2. O valor do plano: número na mesma linha de "Oferta Conjunta Claro MIX" (ex: "310,28"). Formato: vírgula decimal, sem R$.

Também extraia da capa:
3. "cliente": razão social da empresa.
4. "vencimento": data DD/MM/AAAA.

Retorne SOMENTE JSON:
{"cliente": "CONFRUTY ALIMENTOS EIRELI", "vencimento": "24/03/2026", "nome_plano_compartilhado": "Claro Total Compartilhado 150GB", "valor_plano_compartilhado": "310,28"}
"""


def _enriquecer_capa_anthropic(resultado: dict, img_capa, api_key: str) -> None:
    """
    Request extra focado só na capa (1 imagem) para recuperar nome/valor do plano
    quando o request principal não os retornou corretamente.
    Modifica resultado in-place.
    """
    import requests as _req, base64, json as _json
    try:
        buf = io.BytesIO()
        img_capa.save(buf, format="JPEG", quality=85)
        b64 = base64.b64encode(buf.getvalue()).decode()

        content = [
            {"type": "text", "text": _PROMPT_CAPA_COMPARTILHADA},
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}}
        ]
        resp = _req.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json",
                     "anthropic-version": "2023-06-01",
                     "x-api-key": api_key},
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 300,
                  "messages": [{"role": "user", "content": content}]},
            timeout=60
        )
        if resp.status_code != 200:
            return
        texto = "".join(b["text"] for b in resp.json().get("content", []) if b.get("type") == "text")
        texto = re.sub(r"```json|```", "", texto).strip()
        dados = _json.loads(texto)

        # Preenche apenas campos ausentes/inválidos
        if not resultado.get("nome_plano_compartilhado"):
            resultado["nome_plano_compartilhado"] = dados.get("nome_plano_compartilhado", "")
        if not resultado.get("valor_plano_compartilhado") or resultado.get("valor_plano_compartilhado") == "0":
            resultado["valor_plano_compartilhado"] = dados.get("valor_plano_compartilhado", "0")
        if not resultado.get("cliente") or resultado.get("cliente") == "CLIENTE":
            resultado["cliente"] = dados.get("cliente", "CLIENTE")
        if not resultado.get("vencimento"):
            resultado["vencimento"] = dados.get("vencimento", "")
    except Exception:
        pass


def _analisar_compartilhado_com_anthropic(imgs: list) -> dict | None:
    """
    Usa Claude Sonnet para extrair dados de fatura compartilhada em PDF de imagem.
    Se nome/valor do plano não vier no request principal, faz request extra só da capa.
    """
    import requests as _req, base64
    try:
        api_key = None
        try:
            api_key = st.secrets["ANTHROPIC_API_KEY"]
        except Exception:
            pass
        if not api_key:
            return None

        def img_to_b64(img):
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            return base64.b64encode(buf.getvalue()).decode()

        content = [{"type": "text", "text": _PROMPT_FATURA_COMPARTILHADA}]
        for img in imgs:
            content.append({
                "type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": img_to_b64(img)}
            })

        resp = _req.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "anthropic-version": "2023-06-01",
                "x-api-key": api_key,
            },
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 4000,
                  "messages": [{"role": "user", "content": content}]},
            timeout=180
        )
        if resp.status_code != 200:
            return None
        texto = "".join(b["text"] for b in resp.json().get("content", []) if b.get("type") == "text")
        resultado = _parsear_json_compartilhado(texto)

        # Fallback: se nome ou valor do plano vieram vazios, request extra só da capa
        if resultado and (not resultado.get("nome_plano_compartilhado") or
                          not resultado.get("valor_plano_compartilhado") or
                          resultado.get("valor_plano_compartilhado") == "0"):
            _enriquecer_capa_anthropic(resultado, imgs[0], api_key)

        return resultado
    except Exception:
        return None


def _analisar_compartilhado_com_gemini(imgs: list) -> dict | None:
    """
    Usa Gemini para extrair dados de fatura compartilhada em PDF de imagem.
    Mesmo mecanismo de _analisar_com_gemini, com prompt e verificação especializados.
    """
    if not _GEMINI_DISPONIVEL:
        return None

    api_key = None
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        try:
            api_key = st.secrets["GOOGLE_GEMINI_KEY"]
        except Exception:
            pass
    if not api_key:
        return None

    genai.configure(api_key=api_key)
    import time, json as _json

    MODELOS = ["gemini-2.5-flash-lite", "gemini-2.5-flash"]

    def _deduplicar(lista):
        vistos, unicos = set(), []
        for item in lista:
            k = item.get("linha", "")
            if k and k not in vistos:
                vistos.add(k)
                unicos.append(item)
        return unicos

    def _request_com_retry(model, paginas, prompt=None):
        p = prompt or _PROMPT_FATURA_COMPARTILHADA
        for tentativa in range(3):
            try:
                resp = model.generate_content([p] + paginas)
                return resp.text
            except Exception as e:
                err = str(e)
                is_quota = any(x in err for x in ["429", "RESOURCE_EXHAUSTED", "quota"])
                if is_quota and tentativa < 2:
                    time.sleep(15 * (2 ** tentativa))
                elif is_quota:
                    return "QUOTA"
                else:
                    raise
        return None

    def _verificar_internet_compartilhado(model, imgs, linhas_resultado):
        """Segundo request focado só no Subtotal de internet para plano compartilhado."""
        try:
            texto = _request_com_retry(model, imgs, _PROMPT_VERIFICAR_INTERNET_COMPARTILHADO)
            if not texto or texto == "QUOTA":
                return
            texto_clean = re.sub(r"```json|```", "", texto).strip()
            lista = _json.loads(texto_clean)
            if not isinstance(lista, list):
                return
            mapa = {item["linha"]: item["internet_mb"]
                    for item in lista if "linha" in item and "internet_mb" in item}
            for item in linhas_resultado:
                num = item.get("linha", "")
                if num not in mapa:
                    continue
                orig_norm  = _normalizar_internet_mb_ia(item.get("internet_mb", "0"))
                verif_norm = _normalizar_internet_mb_ia(mapa[num])
                orig_f  = float(orig_norm.replace(".", "").replace(",", "."))
                verif_f = float(verif_norm.replace(".", "").replace(",", "."))
                if verif_f > orig_f and (orig_f == 0 or verif_f <= orig_f * 1.5):
                    item["internet_mb"] = mapa[num]
        except Exception:
            pass

    for modelo_nome in MODELOS:
        try:
            model = genai.GenerativeModel(modelo_nome)
            texto = _request_com_retry(model, imgs)
            if texto == "QUOTA":
                st.warning(f"⚠️ Modelo **{modelo_nome}** com quota atingida, tentando alternativa...")
                continue

            resultado = _parsear_json_compartilhado(texto) if texto else None

            if resultado and resultado.get("linhas"):
                resultado["linhas"] = _deduplicar(resultado["linhas"])
                _verificar_internet_compartilhado(model, imgs, resultado["linhas"])
                return resultado

            # Fallback: divide em 2 metades
            metade = len(imgs) // 2
            if metade > 0:
                t1 = _request_com_retry(model, imgs[:metade])
                t2 = _request_com_retry(model, imgs[metade:])
                if t1 == "QUOTA" or t2 == "QUOTA":
                    st.warning(f"⚠️ Modelo **{modelo_nome}** quota atingida, tentando alternativa...")
                    continue
                r1 = _parsear_json_compartilhado(t1) if t1 else None
                r2 = _parsear_json_compartilhado(t2) if t2 else None
                meta = r1 if isinstance(r1, dict) else (r2 if isinstance(r2, dict) else {})
                l1 = (r1 or {}).get("linhas", [])
                l2 = (r2 or {}).get("linhas", [])
                combinado = l1 + l2
                if combinado:
                    linhas_finais = _deduplicar(combinado)
                    _verificar_internet_compartilhado(model, imgs, linhas_finais)
                    return {
                        "tipo": "compartilhado",
                        "cliente": meta.get("cliente", ""),
                        "vencimento": meta.get("vencimento", ""),
                        "nome_plano_compartilhado": meta.get("nome_plano_compartilhado", ""),
                        "valor_plano_compartilhado": meta.get("valor_plano_compartilhado", "0"),
                        "linhas": linhas_finais
                    }

        except Exception as e:
            err = str(e)
            if any(x in err for x in ["429", "RESOURCE_EXHAUSTED", "quota", "not found", "404"]):
                st.warning(f"⚠️ Modelo **{modelo_nome}** indisponível, tentando alternativa...")
                continue
            st.warning(f"⚠️ Gemini ({modelo_nome}): {e}")
            return None

    st.error(
        "❌ **Quota do Gemini esgotada.** Soluções:\n"
        "- Aguarde o reset (~04h Brasília)\n"
        "- Configure `ANTHROPIC_API_KEY` nos Secrets como alternativa\n"
        "- Ative faturamento em [aistudio.google.com](https://aistudio.google.com/app/billing)"
    )
    return None


def _parsear_json_compartilhado(texto: str) -> dict | None:
    """
    Extrai e valida JSON de fatura compartilhada.
    Retorna dict com: tipo, cliente, vencimento, nome_plano_compartilhado,
                      valor_plano_compartilhado, linhas[].
    """
    import json as _json
    try:
        texto = re.sub(r"```json|```", "", texto).strip()
        resultado = _json.loads(texto)
        if not isinstance(resultado, dict):
            return None
        # Valida campos obrigatórios
        if not resultado.get("linhas"):
            return None
        # Garante que tipo está marcado
        resultado["tipo"] = "compartilhado"
        return resultado
    except Exception:
        return None


def analisar_pdf_compartilhado_com_ia(pdf_bytes: bytes) -> dict | None:
    """
    Extrai dados de PDF-imagem de plano compartilhado com IA.
    Usa prompt especializado _PROMPT_FATURA_COMPARTILHADA.
    Ordem: Claude Sonnet (primário) → Gemini (fallback).
    """
    imgs = _converter_paginas(pdf_bytes)
    if not imgs:
        st.error("❌ Não foi possível converter as páginas do PDF em imagens.")
        return None

    _tem_anthropic = False
    try:
        _tem_anthropic = bool(st.secrets["ANTHROPIC_API_KEY"])
    except Exception:
        pass

    if _tem_anthropic:
        resultado = _analisar_compartilhado_com_anthropic(imgs)
        if resultado:
            return resultado
        resultado = _analisar_compartilhado_com_gemini(imgs)
        if resultado:
            return resultado
    else:
        resultado = _analisar_compartilhado_com_gemini(imgs)
        if resultado:
            return resultado

    st.error(
        "❌ Não foi possível analisar o PDF compartilhado com IA.\n"
        "Verifique **GEMINI_API_KEY** ou **ANTHROPIC_API_KEY** nos Secrets."
    )
    return None


def processar_pdf_compartilhado_imagem(resultado_ia: dict, cliente: str, vencimento: str) -> tuple:
    """
    Processa resultado da IA para fatura de plano compartilhado em PDF de imagem.
    Retorna (DataFrame, cliente, vencimento) no mesmo formato que processar_pdf_compartilhado.
    NUNCA altera processar_pdf_compartilhado (digital) nem o fluxo individual.
    """
    nome_plano  = resultado_ia.get("nome_plano_compartilhado", "Claro Total Compartilhado").strip()
    valor_str   = resultado_ia.get("valor_plano_compartilhado", "0")
    valor_plano = to_float(valor_str)
    dados_ia    = resultado_ia.get("linhas", [])

    dados = []
    for item in dados_ia:
        mensalidade_ind = to_float(item.get("mensalidade_individual", "0"))
        valor_pass      = to_float(item.get("valor_passaporte", "0"))
        total_linha     = mensalidade_ind + valor_pass
        internet_raw    = _normalizar_internet_mb_ia(item.get("internet_mb", "0"))

        dados.append({
            "Linha":                  item.get("linha", ""),
            "Internet (MB)":          internet_raw,
            "Internet (MB) fmt":      "",           # preenchido após conversão
            "Pacote de dados":        nome_plano,
            "Mensalidade":            f"R$ {mensalidade_ind:.2f}".replace(".", ","),
            "Passaporte":             item.get("passaporte", "-"),
            "Mensalidade Passaporte": f"R$ {valor_pass:.2f}".replace(".", ",") if valor_pass else "-",
            "Total por linha":        f"R$ {total_linha:.2f}".replace(".", ","),
            "Minutos":                item.get("minutos", "0"),
        })

    df = pd.DataFrame(dados)

    # Converte Internet (MB) para float — mesmo pipeline do fluxo digital
    df["Internet (MB)"] = (
        df["Internet (MB)"]
        .astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df["Internet (MB)"] = pd.to_numeric(df["Internet (MB)"], errors="coerce").fillna(0)
    df["Internet (MB) fmt"] = df["Internet (MB)"].apply(_fmt_mb_display)

    # Classificação de perfil — idêntica a processar_pdf_compartilhado
    franquia_total_mb = extrair_gb_pacote(nome_plano) * 1024

    def classificar_compartilhado(mb):
        if mb == 0:
            return "⚪ Baixo"
        n = max(len(dados), 1)
        media_esperada = franquia_total_mb / n if franquia_total_mb > 0 else 0
        if media_esperada == 0:
            gb = mb / 1024
            if gb < 1:   return "⚪ Baixo"
            if gb < 5:   return "🟡 Médio"
            return "🔴 Alto"
        ratio = mb / media_esperada
        if ratio < 0.3:  return "⚪ Baixo"
        if ratio < 0.8:  return "🟡 Médio"
        return "🔴 Alto"

    df["Perfil"] = df["Internet (MB)"].apply(classificar_compartilhado)

    def em_uso_compartilhado(row):
        minutos_str = str(row["Minutos"]).strip().lower()
        sem_minutos = re.fullmatch(r"0[^\d]*|", minutos_str) is not None
        if row["Internet (MB)"] == 0 and sem_minutos:
            return "Não"
        return "Sim"

    df["Em Uso"] = df.apply(em_uso_compartilhado, axis=1)

    def estrategia_compartilhado(row):
        if row["Em Uso"] == "Não":
            return "⚪ Manter"
        if "Baixo" in row["Perfil"]:
            return "🟡 Sustentar plano"
        if "Médio" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        if "Alto" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        return ""

    df["Estratégia Comercial"] = df.apply(estrategia_compartilhado, axis=1)

    # Linha extra azul — idêntica a processar_pdf_compartilhado
    linha_plano = {
        "Linha":                  "PLANO COMPARTILHADO",
        "Internet (MB)":          0.0,
        "Internet (MB) fmt":      "-",
        "Pacote de dados":        nome_plano,
        "Mensalidade":            f"R$ {valor_plano:.2f}".replace(".", ","),
        "Passaporte":             "-",
        "Mensalidade Passaporte": "-",
        "Total por linha":        f"R$ {valor_plano:.2f}".replace(".", ","),
        "Minutos":                "-",
        "Perfil":                 "-",
        "Em Uso":                 "-",
        "Estratégia Comercial":   "-",
    }
    df = pd.concat([df, pd.DataFrame([linha_plano])], ignore_index=True)

    return df, cliente, vencimento


def _detectar_plano_compartilhado_ia(resultado_ia: dict) -> bool:
    """
    Detecta plano compartilhado em resultado da IA.
    Critério: campo 'tipo' == 'compartilhado' OU presença de 'nome_plano_compartilhado'.
    """
    if resultado_ia.get("tipo") == "compartilhado":
        return True
    if resultado_ia.get("nome_plano_compartilhado"):
        return True
    return False


def _detectar_plano_compartilhado_capa_ia(pdf_bytes: bytes) -> bool:
    """
    Leitura rápida da capa (página 1) com pdfplumber para detectar plano compartilhado
    antes de chamar a IA cara — evita chamar o prompt errado.
    Retorna True se encontrar indicadores de plano compartilhado.
    Fallback seguro: retorna False se não conseguir ler.
    """
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texto_capa = (pdf.pages[0].extract_text() or "") if pdf.pages else ""
        # Para PDF de imagem pdfplumber retorna vazio — usamos detecção fraca
        # mas suficiente para o sinal positivo
        if texto_capa and _detectar_plano_compartilhado(texto_capa):
            return True
    except Exception:
        pass
    return False


def _analisar_pdf_imagem_detectando_tipo(pdf_bytes: bytes) -> dict | None:
    """
    Analisa PDF de imagem detectando automaticamente o tipo de plano.

    Estratégia de detecção (custo zero extra):
    1. Tenta ler capa com pdfplumber (funciona se PDF tiver qualquer texto embutido)
    2. Chama IA com prompt correto:
       - Se compartilhado detectado → _PROMPT_FATURA_COMPARTILHADA
       - Caso contrário → _PROMPT_FATURA (individual)
    3. Após retorno da IA, valida o campo 'tipo' do JSON para confirmação.

    Retorna dict com campo 'tipo': 'compartilhado' | 'individual'
    """
    # Tentativa de detecção via pdfplumber antes de chamar IA
    eh_compartilhado = _detectar_plano_compartilhado_capa_ia(pdf_bytes)

    imgs = _converter_paginas(pdf_bytes)
    if not imgs:
        return None

    _tem_anthropic = False
    try:
        _tem_anthropic = bool(st.secrets["ANTHROPIC_API_KEY"])
    except Exception:
        pass

    if eh_compartilhado:
        # Usa prompt compartilhado diretamente
        if _tem_anthropic:
            resultado = _analisar_compartilhado_com_anthropic(imgs)
            if resultado:
                return resultado
            return _analisar_compartilhado_com_gemini(imgs)
        else:
            return _analisar_compartilhado_com_gemini(imgs)
    else:
        # PDF de imagem sem texto → não sabemos o tipo com certeza.
        # Tenta prompt individual primeiro. Se retornar campo 'tipo'=='compartilhado',
        # reanalisa com prompt compartilhado.
        if _tem_anthropic:
            resultado = _analisar_com_anthropic(imgs)
        else:
            resultado = _analisar_com_gemini(imgs)

        if not resultado:
            return None

        # Verifica se a IA sinalizou compartilhado no retorno
        if _detectar_plano_compartilhado_ia(resultado):
            # Reanalisa com prompt especializado
            if _tem_anthropic:
                r2 = _analisar_compartilhado_com_anthropic(imgs)
                return r2 if r2 else resultado
            else:
                r2 = _analisar_compartilhado_com_gemini(imgs)
                return r2 if r2 else resultado

        return resultado


def _normalizar_internet_mb_ia(valor) -> str:
    """
    Normaliza internet_mb da IA para formato pipeline BR: '7526,866'
    Trata todos os formatos possíveis:
      '14.423.700' → '14423,700' | '14423.700' → '14423,700'
      '14423700'   → '14423,700' | '946,122'   → '946,122'
    """
    s = str(valor).strip()
    if ',' in s:
        return s.replace('.', '')
    if re.search(r'\.\d{3}$', s):
        partes = s.rsplit('.', 1)
        return partes[0].replace('.', '') + ',' + partes[1]
    s_clean = re.sub(r'[^\d]', '', s)
    if not s_clean or s_clean == '0':
        return '0'
    if len(s_clean) > 3:
        return s_clean[:-3] + ',' + s_clean[-3:]
    return '0,' + s_clean.zfill(3)


def _fmt_mb_display(v: float) -> str:
    """Formata float MB para exibição padrão Claro: '14.423,700'"""
    if v == 0:
        return "0"
    inteiro = int(v)
    decimal = round((v - inteiro) * 1000)
    return f"{inteiro:,}".replace(",", ".") + f",{decimal:03d}"


def _validar_pacote_ia(pacote: str, mensalidade_total: str) -> str:
    """
    Normaliza o pacote retornado pela IA.
    Remove nomes de bundle genérico sem GB — esses indicam que o Gemini
    retornou o nome do bundle ao invés do plano individual.
    Preço NÃO é usado como referência — a Claro negocia individualmente.
    """
    if not pacote or pacote == '-':
        return pacote

    # Se tem GB no nome → é o plano individual correto, retorna como está
    if re.search(r'\d+\s*GB', pacote, re.IGNORECASE):
        return pacote

    # Sem GB → nome de bundle genérico (Gemini leu errado)
    # Retorna "-" para indicar que não encontrou o plano individual
    nomes_bundle = ["mix", "oferta conjunta", "bundle", "oferta claro"]
    if any(n in pacote.lower() for n in nomes_bundle):
        return "-"

    # Outros casos sem GB (ex: "Claro Life Ilimitado") → mantém como está
    return pacote



def _parsear_json_ia(texto: str) -> dict | None:
    """
    Extrai JSON da resposta da IA.
    Suporta formato objeto {"cliente":...,"linhas":[...]} e legado array [...].
    Retorna sempre dict com 'cliente', 'vencimento', 'linhas'.
    """
    import json as _json
    try:
        texto = re.sub(r"```json|```", "", texto).strip()
        resultado = _json.loads(texto)
        if isinstance(resultado, dict) and "linhas" in resultado:
            if isinstance(resultado["linhas"], list) and len(resultado["linhas"]) > 0:
                return resultado
        if isinstance(resultado, list) and len(resultado) > 0:
            return {"cliente": "", "vencimento": "", "linhas": resultado}
    except Exception:
        pass
    return None


def _converter_paginas(pdf_bytes: bytes) -> list:
    """Converte PDF em lista de imagens PIL. DPI 200 para melhor leitura de números."""
    try:
        return _pdf2image_convert(pdf_bytes, dpi=200)
    except Exception:
        return []


def _analisar_com_gemini(imgs: list) -> dict | None:
    """
    Usa Gemini para extrair dados das imagens.
    Retorna dict: {"cliente": ..., "vencimento": ..., "linhas": [...]}
    Modelos (free tier): gemini-2.5-flash-lite → gemini-2.5-flash
    Estratégia dupla:
      1. Request principal: extrai todos os campos
      2. Request de verificação: confirma apenas internet_mb via Subtotal
         → usa sempre o maior valor (Subtotal >= linha Internet isolada)
    """
    if not _GEMINI_DISPONIVEL:
        return None

    api_key = None
    try:
        api_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        try:
            api_key = st.secrets["GOOGLE_GEMINI_KEY"]
        except Exception:
            pass
    if not api_key:
        return None

    genai.configure(api_key=api_key)
    import time, json as _json

    MODELOS = ["gemini-2.5-flash-lite", "gemini-2.5-flash"]

    def _deduplicar(lista):
        vistos, unicos = set(), []
        for item in lista:
            k = item.get("linha", "")
            if k and k not in vistos:
                vistos.add(k)
                unicos.append(item)
        return unicos

    def _request_com_retry(model, paginas, prompt=None):
        p = prompt or _PROMPT_FATURA
        for tentativa in range(3):
            try:
                resp = model.generate_content([p] + paginas)
                return resp.text
            except Exception as e:
                err = str(e)
                is_quota = any(x in err for x in ["429", "RESOURCE_EXHAUSTED", "quota"])
                if is_quota and tentativa < 2:
                    time.sleep(15 * (2 ** tentativa))
                elif is_quota:
                    return "QUOTA"
                else:
                    raise
        return None

    def _verificar_internet(model, imgs, linhas_resultado):
        """
        Segundo request focado só no Subtotal de Internet.
        Corrige silenciosamente se o Gemini pegou só 'Internet' sem 'meses anteriores'.
        Só substitui se o valor verificado for maior E plausível (≤ 3x o original),
        evitando contaminação entre linhas quando o Gemini associa Subtotal errado.
        """
        try:
            texto = _request_com_retry(model, imgs, _PROMPT_VERIFICAR_INTERNET)
            if not texto or texto == "QUOTA":
                return
            texto_clean = re.sub(r"```json|```", "", texto).strip()
            lista = _json.loads(texto_clean)
            if not isinstance(lista, list):
                return
            mapa = {item["linha"]: item["internet_mb"]
                    for item in lista if "linha" in item and "internet_mb" in item}
            for item in linhas_resultado:
                num = item.get("linha", "")
                if num not in mapa:
                    continue
                orig_norm  = _normalizar_internet_mb_ia(item.get("internet_mb", "0"))
                verif_norm = _normalizar_internet_mb_ia(mapa[num])
                orig_f  = float(orig_norm.replace(".", "").replace(",", "."))
                verif_f = float(verif_norm.replace(".", "").replace(",", "."))
                # Só substitui se: (1) verificado é maior E (2) não mais que 1.5x o original
                # 1.5x cobre "meses anteriores" legítimos (tipicamente 1.0x–1.3x)
                # mas bloqueia contaminação entre linhas (tipicamente 1.9x ou mais)
                if verif_f > orig_f and (orig_f == 0 or verif_f <= orig_f * 1.5):
                    item["internet_mb"] = mapa[num]
        except Exception:
            pass  # falha silenciosa — mantém resultado original

    for modelo_nome in MODELOS:
        try:
            model = genai.GenerativeModel(modelo_nome)

            # Request principal
            texto = _request_com_retry(model, imgs)
            if texto == "QUOTA":
                st.warning(f"⚠️ Modelo **{modelo_nome}** com quota atingida, tentando alternativa...")
                continue

            resultado = _parsear_json_ia(texto) if texto else None

            if resultado and resultado.get("linhas"):
                resultado["linhas"] = _deduplicar(resultado["linhas"])
                # Verificação de internet_mb (segundo request)
                _verificar_internet(model, imgs, resultado["linhas"])
                return resultado

            # Fallback: divide em 2 metades se retornou vazio
            metade = len(imgs) // 2
            if metade > 0:
                t1 = _request_com_retry(model, imgs[:metade])
                t2 = _request_com_retry(model, imgs[metade:])
                if t1 == "QUOTA" or t2 == "QUOTA":
                    st.warning(f"⚠️ Modelo **{modelo_nome}** quota atingida, tentando alternativa...")
                    continue
                r1 = _parsear_json_ia(t1) if t1 else None
                r2 = _parsear_json_ia(t2) if t2 else None
                meta = r1 if isinstance(r1, dict) else (r2 if isinstance(r2, dict) else {})
                l1 = (r1 or {}).get("linhas", [])
                l2 = (r2 or {}).get("linhas", [])
                combinado = l1 + l2
                if combinado:
                    linhas_finais = _deduplicar(combinado)
                    _verificar_internet(model, imgs, linhas_finais)
                    return {
                        "cliente": meta.get("cliente", ""),
                        "vencimento": meta.get("vencimento", ""),
                        "linhas": linhas_finais
                    }

        except Exception as e:
            err = str(e)
            if any(x in err for x in ["429", "RESOURCE_EXHAUSTED", "quota", "not found", "404"]):
                st.warning(f"⚠️ Modelo **{modelo_nome}** indisponível, tentando alternativa...")
                continue
            st.warning(f"⚠️ Gemini ({modelo_nome}): {e}")
            return None

    st.error(
        "❌ **Quota do Gemini esgotada.** Soluções:\n"
        "- Aguarde o reset (~04h Brasília)\n"
        "- Configure `ANTHROPIC_API_KEY` nos Secrets como alternativa\n"
        "- Ative faturamento em [aistudio.google.com](https://aistudio.google.com/app/billing)"
    )
    return None


def _analisar_com_anthropic(imgs: list) -> dict | None:
    """Usa Claude Sonnet (pago) como fallback se Gemini não estiver configurado."""
    import requests as _req, base64
    try:
        api_key = None
        try:
            api_key = st.secrets["ANTHROPIC_API_KEY"]
        except Exception:
            pass
        if not api_key:
            return None

        def img_to_b64(img):
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            return base64.b64encode(buf.getvalue()).decode()

        content = [{"type": "text", "text": _PROMPT_FATURA}]
        for img in imgs:
            content.append({
                "type": "image",
                "source": {"type": "base64", "media_type": "image/jpeg", "data": img_to_b64(img)}
            })

        resp = _req.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type": "application/json",
                "anthropic-version": "2023-06-01",
                "x-api-key": api_key,
            },
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 2000,
                  "messages": [{"role": "user", "content": content}]},
            timeout=120
        )
        if resp.status_code != 200:
            return None
        texto = "".join(b["text"] for b in resp.json().get("content", []) if b.get("type") == "text")
        return _parsear_json_ia(texto)
    except Exception:
        return None


def analisar_pdf_imagem_com_ia(pdf_bytes: bytes) -> dict | None:
    """
    Extrai dados de PDF-imagem com IA.

    Ordem de preferência:
      1. Claude Sonnet (ANTHROPIC_API_KEY) — mais preciso, ~$0,03/fatura
      2. Gemini gratuito (GEMINI_API_KEY)  — gratuito, adequado para faturas simples

    Se ANTHROPIC_API_KEY estiver configurado, usa Claude primeiro.
    Se não estiver, usa Gemini.
    """
    imgs = _converter_paginas(pdf_bytes)
    if not imgs:
        st.error("❌ Não foi possível converter as páginas do PDF em imagens.")
        return None

    # Verifica se Anthropic está configurado
    _tem_anthropic = False
    try:
        _tem_anthropic = bool(st.secrets["ANTHROPIC_API_KEY"])
    except Exception:
        pass

    if _tem_anthropic:
        # Claude primeiro — mais preciso para faturas com muitas linhas
        resultado = _analisar_com_anthropic(imgs)
        if resultado:
            return resultado
        # Fallback para Gemini se Claude falhar
        resultado = _analisar_com_gemini(imgs)
        if resultado:
            return resultado
    else:
        # Gemini primeiro (gratuito)
        resultado = _analisar_com_gemini(imgs)
        if resultado:
            return resultado

    st.error(
        "❌ Não foi possível analisar o PDF com IA.\n"
        "Verifique **GEMINI_API_KEY** nos Secrets do Streamlit Cloud.\n"
        "Para maior precisão em faturas com muitas linhas, configure também **ANTHROPIC_API_KEY**."
    )
    return None


def processar_pdf(file):
    """
    Processa PDF da Claro com 2 estratégias:
    1. PDF digital  → pdfplumber + regex
    2. PDF imagem   → IA (Gemini/Anthropic) com dupla verificação
    """
    placeholder = st.empty()
    file.seek(0)
    pdf_bytes = file.read()

    # Detectar tipo: pelo menos 100 chars nas primeiras 3 páginas = digital
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total_paginas = len(pdf.pages)
        chars_extraidos = sum(
            len((p.extract_text() or "").strip())
            for p in pdf.pages[:3]
        )
    eh_imagem = (chars_extraidos < 100)

    if not eh_imagem:
        # ── PDF digital: extração por pdfplumber + regex ──
        texto = ""
        progresso = st.progress(0)
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for i, page in enumerate(pdf.pages):
                placeholder.text(f"📄 Processando página {i+1} de {total_paginas}...")
                t = page.extract_text()
                if t and t.strip():
                    texto += t + "\n"
                progresso.progress((i + 1) / total_paginas)
        placeholder.text("🔎 Extraindo dados...")
        progresso.empty()
        placeholder.empty()

        cliente    = extrair_cliente(texto)
        vencimento = extrair_vencimento(texto)

        # ── Detecção: plano compartilhado ou individual? ──
        if _detectar_plano_compartilhado(texto):
            return processar_pdf_compartilhado(texto, cliente, vencimento)

        # ── Fluxo individual (original, intocado) ──
        linhas     = extrair_linhas(texto)
        blocos     = extrair_blocos_por_linha(texto)
        mensalidades = extrair_mensalidades(blocos)
        detalhamento = extrair_detalhamento(blocos)
        pacotes    = extrair_pacote_e_passaporte(blocos)

        dados = []
        for linha in linhas:
            total       = to_float(mensalidades.get(linha, "0"))
            valor_pass  = to_float(pacotes.get(linha, {}).get("Valor Passaporte", "0"))
            valor_plano = total - valor_pass
            dados.append({
                "Linha":                  linha,
                "Internet (MB)":          detalhamento.get(linha, {}).get("Internet (MB)", "0"),
                "Pacote de dados":        pacotes.get(linha, {}).get("Pacote", "-"),
                "Mensalidade":            f"R$ {valor_plano:.2f}".replace(".", ","),
                "Passaporte":             pacotes.get(linha, {}).get("Passaporte", "-"),
                "Mensalidade Passaporte": f"R$ {valor_pass:.2f}".replace(".", ",") if valor_pass else "-",
                "Total por linha":        f"R$ {total:.2f}".replace(".", ","),
                "Minutos":                detalhamento.get(linha, {}).get("Minutos", "0"),
            })

    else:
        # ── PDF imagem: IA com detecção de tipo e dupla verificação ──
        if not _PDF2IMAGE_DISPONIVEL:
            raise ValueError("pdf2image não instalado. Adicione ao requirements.txt e poppler-utils ao packages.txt.")

        placeholder.text("🖼️ PDF de imagem — analisando com IA (pode levar ~30s)...")
        progresso = st.progress(0.1)

        resultado_ia = _analisar_pdf_imagem_detectando_tipo(pdf_bytes)

        progresso.progress(1.0)
        progresso.empty()
        placeholder.empty()

        if not resultado_ia:
            raise ValueError("Não foi possível extrair dados com IA. Verifique a chave GEMINI_API_KEY nos Secrets.")

        cliente    = resultado_ia.get("cliente", "").strip().upper() or "CLIENTE"
        vencimento = resultado_ia.get("vencimento", "").strip()

        # Fallback Google Vision para cliente/vencimento se IA não retornou
        if cliente == "CLIENTE" or not vencimento:
            try:
                from google.cloud import vision as _v
                from google.oauth2 import service_account as _sa
                _creds = _sa.Credentials.from_service_account_info(st.secrets["GOOGLE_CREDENTIALS"])
                _cli = _v.ImageAnnotatorClient(credentials=_creds)
                _imgs_capa = _pdf2image_convert(pdf_bytes, dpi=120, first_page=1, last_page=1)
                _buf = io.BytesIO(); _imgs_capa[0].save(_buf, format="PNG")
                _resp = _cli.document_text_detection(image=_v.Image(content=_buf.getvalue()))
                _t = _resp.full_text_annotation.text or ""
                c = extrair_cliente(_t); v = extrair_vencimento(_t)
                if c != "CLIENTE" and cliente == "CLIENTE": cliente = c
                if v and not vencimento: vencimento = v
            except Exception:
                pass

        # ── Desvio para processador de plano compartilhado (imagem) ──
        # Não toca em processar_pdf_compartilhado (digital) nem no fluxo individual.
        if _detectar_plano_compartilhado_ia(resultado_ia):
            return processar_pdf_compartilhado_imagem(resultado_ia, cliente, vencimento)

        # ── Fluxo individual imagem (original, intocado) ──
        dados_ia = resultado_ia.get("linhas", [])

        dados = []
        for item in dados_ia:
            total       = to_float(item.get("mensalidade_total", "0"))
            valor_pass  = to_float(item.get("valor_passaporte", "0"))
            valor_plano = total - valor_pass
            pacote_ok   = _validar_pacote_ia(item.get("pacote", "-"), item.get("mensalidade_total", "0"))
            dados.append({
                "Linha":                  item.get("linha", ""),
                "Internet (MB)":          _normalizar_internet_mb_ia(item.get("internet_mb", "0")),
                "Pacote de dados":        pacote_ok,
                "Mensalidade":            f"R$ {valor_plano:.2f}".replace(".", ","),
                "Passaporte":             item.get("passaporte", "-"),
                "Mensalidade Passaporte": f"R$ {valor_pass:.2f}".replace(".", ",") if valor_pass else "-",
                "Total por linha":        f"R$ {total:.2f}".replace(".", ","),
                "Minutos":                item.get("minutos", "0"),
            })

    df = pd.DataFrame(dados)

    df["Internet (MB)"] = (
        df["Internet (MB)"]
        .astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    df["Internet (MB)"] = pd.to_numeric(df["Internet (MB)"], errors="coerce").fillna(0)

    # Coluna formatada para exibição: '14.423,700'
    df["Internet (MB) fmt"] = df["Internet (MB)"].apply(_fmt_mb_display)

    def classificar(x):
        if x > 10000:
            return "🔴 Alto"
        elif x > 3000:
            return "🟡 Médio"
        else:
            return "⚪ Baixo"

    df["Perfil"] = df["Internet (MB)"].apply(classificar)

    def em_uso(row):
        minutos_str = str(row["Minutos"]).strip().lower()
        sem_minutos = re.fullmatch(r"0[^\d]*|", minutos_str) is not None
        if row["Internet (MB)"] == 0 and sem_minutos:
            return "Não"
        return "Sim"

    df["Em Uso"] = df.apply(em_uso, axis=1)

    def estrategia(row):
        if row["Em Uso"] == "Não":
            return "⚪ Manter"
        pacote_gb = extrair_gb_pacote(row["Pacote de dados"])
        uso_gb = row["Internet (MB)"] / 1024 if row["Internet (MB)"] else 0
        if pacote_gb > 0 and uso_gb >= pacote_gb * 0.9:
            return "🔵 Upsell → Aumento recomendado"
        if "Baixo" in row["Perfil"]:
            return "🟡 Sustentar plano"
        if "Médio" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        if "Alto" in row["Perfil"]:
            return "🟢 Bem dimensionado"
        return ""

    df["Estratégia Comercial"] = df.apply(estrategia, axis=1)

    return df, cliente, vencimento

def gerar_excel(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalhamento"

    df_reset = df.reset_index(drop=True)

    for r in dataframe_to_rows(df_reset, index=False, header=True):
        ws.append(r)

    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="333333", fill_type="solid")
    zebra       = PatternFill(start_color="F2F2F2", fill_type="solid")
    vermelho    = PatternFill(start_color="FF4C4C", fill_type="solid")
    verde       = PatternFill(start_color="C6EFCE", fill_type="solid")
    amarelo     = PatternFill(start_color="FFF3B0", fill_type="solid")
    azul        = PatternFill(start_color="BDD7EE", fill_type="solid")
    cinza       = PatternFill(start_color="D9D9D9", fill_type="solid")

    headers = [cell.value for cell in ws[1]]
    col_idx = {v: i for i, v in enumerate(headers)}

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    azul_plano  = PatternFill(start_color="BDD7EE", fill_type="solid")
    fonte_plano = Font(bold=True, color="1F4E79")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for j, cell in enumerate(row):
            coluna = headers[j]
            if coluna in ("Perfil", "Estratégia Comercial"):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

        # Linha do plano compartilhado — destaque azul
        linha_val = str(row[col_idx["Linha"]].value) if "Linha" in col_idx else ""
        if linha_val == "PLANO COMPARTILHADO":
            for cell in row:
                cell.fill = azul_plano
                cell.font = fonte_plano
            continue  # não aplica zebra nem coloração de perfil

        if i % 2 == 0:
            for cell in row:
                cell.fill = zebra

        perfil     = str(row[col_idx["Perfil"]].value) if "Perfil" in col_idx else ""
        uso        = str(row[col_idx["Em Uso"]].value) if "Em Uso" in col_idx else ""
        estrategia_val = str(row[col_idx["Estratégia Comercial"]].value) if "Estratégia Comercial" in col_idx else ""

        if "Alto" in perfil:
            row[col_idx["Perfil"]].fill = vermelho
        elif "Médio" in perfil:
            row[col_idx["Perfil"]].fill = amarelo

        if "Em Uso" in col_idx:
            if uso == "Não":
                row[col_idx["Em Uso"]].fill = vermelho
            else:
                row[col_idx["Em Uso"]].fill = verde

        if "Estratégia Comercial" in col_idx:
            if "Manter" in estrategia_val:
                row[col_idx["Estratégia Comercial"]].fill = cinza
            elif "Sustentar" in estrategia_val:
                row[col_idx["Estratégia Comercial"]].fill = amarelo
            elif "Bem dimensionado" in estrategia_val:
                row[col_idx["Estratégia Comercial"]].fill = verde
            elif "Upsell" in estrategia_val:
                row[col_idx["Estratégia Comercial"]].fill = azul

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ===== EXECUÇÃO =====
# ===== EXECUÇÃO =====
# Usa session_state para evitar re-análise ao clicar em botões (ex: download Excel)
# A análise só é refeita quando os arquivos enviados mudam.

# Gera uma chave única baseada nos arquivos carregados
_arquivos_key = tuple(
    (f.name, f.size) for f in uploaded_files
) if uploaded_files else ()

# Se os arquivos mudaram, limpa o cache
if st.session_state.get("_arquivos_key") != _arquivos_key:
    st.session_state["_arquivos_key"] = _arquivos_key
    st.session_state["_df_total"] = None
    st.session_state["_cliente_nome"] = "CLIENTE"
    st.session_state["_vencimento_fatura"] = ""

if uploaded_files:
    # Só processa se ainda não tiver resultado em cache
    if st.session_state.get("_df_total") is None:
        df_total = pd.DataFrame()
        cliente_nome = "CLIENTE"
        vencimento_fatura = ""

        progress = st.progress(0)
        total_files = len(uploaded_files)

        for i, file in enumerate(uploaded_files):
            try:
                with st.spinner(f"Processando {file.name}..."):
                    df, cliente, vencimento = processar_pdf(file)
                    df_total = pd.concat([df_total, df], ignore_index=True)
                    cliente_nome = cliente
                    vencimento_fatura = vencimento
            except Exception as e:
                st.error(f"❌ Erro ao processar **{file.name}**: {e}")
                continue

            progress.progress((i + 1) / total_files)

        # Salva resultado no session_state
        st.session_state["_df_total"] = df_total
        st.session_state["_cliente_nome"] = cliente_nome
        st.session_state["_vencimento_fatura"] = vencimento_fatura

    # Recupera resultado do cache (sem re-processar)
    df_total = st.session_state["_df_total"]
    cliente_nome = st.session_state["_cliente_nome"]
    vencimento_fatura = st.session_state["_vencimento_fatura"]

    if df_total is not None and not df_total.empty:
        st.markdown('<div class="tt-divider"></div>', unsafe_allow_html=True)
        st.markdown('<p class="tt-section-title">📊 Resumo da Fatura</p>', unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)

        # Exclui linha especial "PLANO COMPARTILHADO" dos cálculos de resumo
        df_linhas = df_total[df_total["Linha"] != "PLANO COMPARTILHADO"]

        total_linhas = len(df_linhas)
        em_uso       = (df_linhas["Em Uso"] == "Sim").sum()
        inativas     = total_linhas - em_uso
        total_gb     = df_linhas["Internet (MB)"].sum() / 1024
        media_gb     = total_gb / total_linhas if total_linhas else 0

        col1.metric("Total de Linhas", total_linhas)
        col2.metric("Linhas Ativas", em_uso, delta=f"{inativas} inativas" if inativas else None,
                    delta_color="inverse")
        col3.metric("Consumo Total", f"{round(total_gb, 1)} GB")
        col4.metric("Média por Linha", f"{round(media_gb, 1)} GB")

        st.markdown('<div class="tt-divider"></div>', unsafe_allow_html=True)
        st.markdown('<p class="tt-section-title">📋 Detalhamento por Linha</p>', unsafe_allow_html=True)

        # Exibição com MB formatado '14.423,700'
        df_display = df_total.copy()
        if "Internet (MB) fmt" in df_display.columns:
            df_display["Internet (MB)"] = df_display["Internet (MB) fmt"]
            df_display = df_display.drop(columns=["Internet (MB) fmt"])

        st.dataframe(
            df_display,
            use_container_width=True,
            height=min(600, 60 + len(df_display) * 38),
            hide_index=True,
        )

        st.markdown('<div class="tt-divider"></div>', unsafe_allow_html=True)

        col_info, col_btn = st.columns([3, 1])
        with col_info:
            st.markdown(f"""
            <div style="padding:14px 18px;background:rgba(16,185,129,0.06);
                        border:1px solid rgba(16,185,129,0.15);border-radius:12px;">
                <div style="font-size:0.72rem;color:#64748b;text-transform:uppercase;
                            letter-spacing:0.08em;font-weight:600;margin-bottom:4px;">
                    Relatório pronto
                </div>
                <div style="font-size:0.9rem;color:#cbd5e1;">
                    <strong style="color:#10b981;">{total_linhas}</strong> linhas ·
                    cliente <strong style="color:#e2e8f0;">{cliente_nome}</strong> ·
                    vencimento <strong style="color:#e2e8f0;">{vencimento_fatura or '—'}</strong>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with col_btn:
            excel = gerar_excel(df_total.drop(columns=["Internet (MB) fmt"], errors="ignore"))
            nome_arquivo = (
                f"Analise_Target_{cliente_nome}_{vencimento_fatura}.xlsx"
                if vencimento_fatura
                else f"Analise_Target_{cliente_nome}.xlsx"
            )
            st.download_button(
                "📥  Baixar Relatório Excel",
                data=excel,
                file_name=nome_arquivo,
                use_container_width=True,
            )
